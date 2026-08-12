import os, sys, datetime
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

import_base = r"C:\Ersen\Projects_2025\Uno_ERP\Uno_E2E_Tests\20260731_import"
f1_dir = os.path.join(import_base, r"FromUno\3-10 Temmuz, Turgut")
f2_dir = os.path.join(import_base, r"FromUno\5-12 Temmuz, Levent")

print("Regenerating import files with CORRECT Master Data Metadata template structure and filtering zero-sales passengers...")

# Delete previous files if present
for f_dir in [f1_dir, f2_dir]:
    for fname in os.listdir(f_dir):
        if fname.endswith("_importMetadata.xlsx") or fname.endswith("_importrooming.xlsx") or fname.endswith("_importSales.xlsx"):
            try:
                os.remove(os.path.join(f_dir, fname))
                print(f"Deleted: {fname}")
            except Exception as e:
                print(f"Error deleting {fname}: {e}")

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

def parse_rooming_strict(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.worksheets[0]
    pax_list = []
    current_room = ""
    
    for r in range(2, sheet.max_row + 1):
        r_type = sheet.cell(r, 1).value
        gender = clean_str(sheet.cell(r, 2).value)
        first = clean_str(sheet.cell(r, 3).value)
        last = clean_str(sheet.cell(r, 4).value)
        bday = sheet.cell(r, 5).value
        
        if r_type is not None and str(r_type).strip() != "":
            raw_room = str(r_type).strip().upper()
            if "SINGLE" in raw_room:
                current_room = "Single"
            elif "TWIN" in raw_room:
                current_room = "Twin"
            elif "TRIPL" in raw_room or "TRIPLE" in raw_room:
                current_room = "Triple"
            else:
                current_room = "Double"
                
        if first != "" or last != "":
            pax_type = "Adult"
            if gender == "CHD" or (bday and isinstance(bday, str) and len(bday) > 4 and int(bday.split('.')[-1]) >= 2014):
                pax_type = "Children"
            elif gender == "INF" or (bday and isinstance(bday, str) and len(bday) > 4 and int(bday.split('.')[-1]) >= 2024):
                pax_type = "Infant"
                
            pax_list.append({
                "first": first,
                "last": last,
                "gender": gender if gender else None,
                "room_type": current_room,
                "pax_type": pax_type,
                "bday": bday,
                "tc": None,
                "phone": None,
                "passport": None,
                "passport_type": None,
                "address": None
            })
    return pax_list

def parse_sales_strict(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.worksheets[0]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    
    exc_cols = []
    for c_idx, h in enumerate(headers, 1):
        if h is not None and c_idx >= 6:
            h_str = str(h).replace('\n', ' ').strip()
            parts = h_str.split()
            price = None
            for p in parts:
                if p.isdigit():
                    price = int(p)
                    break
            exc_cols.append({
                "col_idx": c_idx,
                "code": h_str,
                "name": h_str,
                "price": price
            })
            
    sales_data = {}
    for r in range(2, sheet.max_row + 1):
        first = clean_str(sheet.cell(r, 2).value)
        last = clean_str(sheet.cell(r, 3).value)
        if first != "" or last != "":
            key = (first.upper(), last.upper())
            sales_data[key] = {}
            for ec in exc_cols:
                val = sheet.cell(r, ec["col_idx"]).value
                bought = False
                if val is not None:
                    v_str = str(val).strip().upper()
                    if v_str not in ["", "0", "FALSE", "NONE", "-"]:
                        bought = True
                sales_data[key][ec["code"]] = bought
                
    return exc_cols, sales_data

def parse_fatura_handling(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.worksheets[0]
    unit_price = None
    for r in range(1, sheet.max_row + 1):
        v = sheet.cell(r, 1).value
        if v and "Handling" in str(v):
            unit_price = sheet.cell(r, 7).value
            break
    return unit_price


# =========================================================
# 1. CREATE MASTER DATA METADATA WORKBOOK (_importMetadata.xlsx)
# Matching screenshot SS1 & MasterData_Import_Template.xlsx / uno_import_template.xlsx
# =========================================================
def create_master_data_metadata_wb(group_info, pax_list, exc_cols):
    wb = openpyxl.Workbook()
    
    # 1. Clients
    ws_clients = wb.active
    ws_clients.title = "Clients"
    ws_clients.append(["Name", "Location"])
    ws_clients.append(["WTATIL Seyahat Tur Acentalıgı Ltd.Stı.", "Istanbul / Turkey"])
    
    # 2. Projects
    ws_projects = wb.create_sheet(title="Projects")
    ws_projects.append(["ProjectCode", "Description", "ClientName", "StartDate", "EndDate", "EstimatedBudget"])
    ws_projects.append([group_info["project_code"], group_info["description"], "WTatil", group_info["start_date"], group_info["end_date"], None])
    
    # 3. Tours
    ws_tours = wb.create_sheet(title="Tours")
    ws_tours.append(["TourCode", "ProjectCode", "Destination", "ArrivalDate", "EndDate", "Adults", "Children", "Infants", "Pax", "Adult Rate", "Child Rate", "Infant Rate", "Base Fee", "Total Fee"])
    adults = sum(1 for p in pax_list if p["pax_type"] == "Adult")
    children = sum(1 for p in pax_list if p["pax_type"] == "Children")
    infants = sum(1 for p in pax_list if p["pax_type"] == "Infant")
    ws_tours.append([group_info["tour_code"], group_info["project_code"], "Budapest-Vienna-Prague", group_info["start_date"], group_info["end_date"], adults, children, infants, len(pax_list), None, None, None, None, None])
    
    # 4. Hotels (Headers only)
    ws_hotels = wb.create_sheet(title="Hotels")
    ws_hotels.append(["Name", "Location", "StarRating", "ContactInfo", "SingleRate", "DoubleRate", "TwinRate", "TripleRate"])
    
    # 5. Guides
    ws_guides = wb.create_sheet(title="Guides")
    ws_guides.append(["Name", "Language", "PhoneNumber", "DailyRate"])
    ws_guides.append([group_info["guide_name"], "Turkish / English", None, None])
    
    # 6. TransportCompanies (Headers only)
    ws_tc = wb.create_sheet(title="TransportCompanies")
    ws_tc.append(["Name", "ContactInfo", "FleetSize", "DailyRate"])
    
    # 7. Drivers (Headers only)
    ws_drv = wb.create_sheet(title="Drivers")
    ws_drv.append(["Name", "PhoneNumber", "TransportCompany", "DailyRate"])
    
    # 8. Excursions (List of optional excursions from tour list)
    ws_exc = wb.create_sheet(title="Excursions")
    ws_exc.append(["Name", "Type", "CostPerPax", "SalePrice"])
    for ec in exc_cols:
        ws_exc.append([ec["name"], "Optional Excursion", None, ec["price"]])
        
    # 9. Bookings (Headers only)
    ws_bk = wb.create_sheet(title="Bookings")
    ws_bk.append(["TourCode", "ClientName", "BookingDate", "Status", "ServiceType", "TotalAmount"])
    
    # 10. TourServices (Headers only)
    ws_ts = wb.create_sheet(title="TourServices")
    ws_ts.append(["TourCode", "Category", "Description", "Quantity", "UnitPrice", "RoomType", "RoomCount", "HotelName", "GuideName", "DriverName", "TransportCompany", "ExcursionName"])
    
    # 11. Vendors (Headers only)
    ws_vnd = wb.create_sheet(title="Vendors")
    ws_vnd.append(["Name", "Contact", "ServiceType"])
    
    # 12. ExcursionVendors (Headers only)
    ws_ev = wb.create_sheet(title="ExcursionVendors")
    ws_ev.append(["ExcursionName", "VendorName", "Cost", "Notes"])
    
    return wb

# =========================================================
# 2. CREATE ROOMING WORKBOOK (_importrooming.xlsx)
# Matching TourImportTemplate.xlsx & PRJ-BVP2_PVB07092026_rooming.xlsx
# =========================================================
def create_rooming_manifest_wb(pax_list):
    wb = openpyxl.Workbook()
    
    # 1. Rooming Sheet
    ws_r = wb.active
    ws_r.title = "Rooming"
    ws_r.append(["Yolcu Adı", "Yolcu Soyadı", "Cinsiyet", "RoomType", "Pax Type", "T.C. Kimlik No", "Doğum Tarihi", "Telefon", "Pasaport No", "Pasaport type", "Visa No", "Address"])
    for p in pax_list:
        ws_r.append([p["first"], p["last"], p["gender"], p["room_type"], p["pax_type"], p["tc"], p["bday"], p["phone"], p["passport"], p["passport_type"], None, p["address"]])
        
    # 2. Hotels Sheet (Headers only)
    ws_h = wb.create_sheet(title="Hotels")
    ws_h.append(["Hotel Name", "Status", "Contact", "Check-in", "Check-out", "Double", "Single", "Triple", "Twin"])
    
    # 3. Flights Sheet (Headers only)
    ws_f = wb.create_sheet(title="Flights")
    ws_f.append(["Flight No", "Origin", "Destination", "Date (dd.MM.yyyy)", "Time"])
    
    # 4. Sheet1 (Pax Summary Pivot Table)
    ws_s1 = wb.create_sheet(title="Sheet1")
    ws_s1.append(["Row Labels", "Count of Yolcu Adı"])
    adults = sum(1 for p in pax_list if p["pax_type"] == "Adult")
    children = sum(1 for p in pax_list if p["pax_type"] == "Children")
    infants = sum(1 for p in pax_list if p["pax_type"] == "Infant")
    ws_s1.append(["Adult", adults])
    ws_s1.append(["Children", children])
    ws_s1.append(["Infant", infants])
    ws_s1.append(["Grand Total", len(pax_list)])
    
    return wb

# =========================================================
# 3. CREATE SALES WORKBOOK (_importSales.xlsx)
# Matching SalesImportTemplate_v2.xlsx
# STRICT: FILTER PASSENGERS WITH 0 EXCURSION SALES
# =========================================================
def create_sales_matrix_wb(pax_list, exc_cols, sales_data, unit_price):
    wb = openpyxl.Workbook()
    
    # 1. ExcusionSales Sheet
    ws_es = wb.active
    ws_es.title = "ExcusionSales"
    
    # Row 1: Dates (BLANK if not in source)
    ws_es.append(["Dates"] + [None for _ in exc_cols])
    # Row 2: Prices
    ws_es.append(["Prices"] + [ec["price"] for ec in exc_cols])
    # Row 3: Code
    ws_es.append(["Code"] + [ec["code"] for ec in exc_cols])
    # Row 4: Excursion Name
    ws_es.append(["Passenger Name"] + [ec["name"] for ec in exc_cols])
    
    # Row 5+: Passenger Excursion Matrix (STRICTLY INCLUDE ONLY PASSENGERS WITH AT LEAST ONE SALE)
    included_count = 0
    for p in pax_list:
        full_name = f"{p['first']} {p['last']}".strip()
        key = (p["first"].upper(), p["last"].upper())
        p_sales = sales_data.get(key, {})
        
        # Check if passenger bought AT LEAST ONE excursion
        has_any_sale = any(p_sales.get(ec["code"], False) for ec in exc_cols)
        
        if has_any_sale:
            row_vals = [full_name]
            for ec in exc_cols:
                bought = p_sales.get(ec["code"], False)
                row_vals.append(True if bought else False)
            ws_es.append(row_vals)
            included_count += 1
            
    print(f"   -> ExcusionSales: Included {included_count} passengers with active sales (Skipped {len(pax_list) - included_count} passengers with 0 sales).")
    
    # 2. BaseServices Sheet
    ws_bs = wb.create_sheet(title="BaseServices")
    ws_bs.append(["Base Service", "Revenue", "Expense", "Other", "per/Pax", "UnitPrice", "Adult", "Children", "Infant", "Total"])
    
    adults = sum(1 for p in pax_list if p["pax_type"] == "Adult")
    children = sum(1 for p in pax_list if p["pax_type"] == "Children")
    infants = sum(1 for p in pax_list if p["pax_type"] == "Infant")
    
    # Row 2: Agency Fee (from fatura unit price if available)
    ws_bs.append(["Agency Fee", True, False, False, True, unit_price, adults, children, infants, "=(G2+ (H2*5) + (I2*0)) * F2"])
    # Row 3: CityTax
    ws_bs.append(["CityTax", True, False, False, True, None, adults, children, infants, "=(G3+H3+I3) *F3"])
    
    return wb

# =========================================================
# GENERATE FOR GROUP 1 (3-10 Temmuz, Turgut)
# =========================================================
print("\n--- Generating Group 1 Files (3-10 Temmuz, Turgut) ---")
pax1 = parse_rooming_strict(os.path.join(f1_dir, "03_1007 BUD-PRG RL.xlsx"))
exc1, sales1 = parse_sales_strict(os.path.join(f1_dir, "0307_1007 tur listesi.xlsx"))
price1 = parse_fatura_handling(os.path.join(f1_dir, "03_10.07 fatura.xlsx"))

g1_info = {
    "project_code": "PRJ-BVP",
    "description": "3-10 Temmuz Turgut Orta Avrupa Turu",
    "tour_code": "BVP03072026",
    "guide_name": "Turgut",
    "start_date": "03.07.2026",
    "end_date": "10.07.2026"
}

wb_meta1 = create_master_data_metadata_wb(g1_info, pax1, exc1)
wb_meta1.save(os.path.join(f1_dir, "3-10_Temmuz_Turgut_importMetadata.xlsx"))
print("Saved Master Data Metadata Template: 3-10_Temmuz_Turgut_importMetadata.xlsx")

wb_room1 = create_rooming_manifest_wb(pax1)
wb_room1.save(os.path.join(f1_dir, "3-10_Temmuz_Turgut_importrooming.xlsx"))
print("Saved Rooming Manifest Template: 3-10_Temmuz_Turgut_importrooming.xlsx")

wb_sales1 = create_sales_matrix_wb(pax1, exc1, sales1, price1)
wb_sales1.save(os.path.join(f1_dir, "3-10_Temmuz_Turgut_importSales.xlsx"))
print("Saved Sales Matrix Template: 3-10_Temmuz_Turgut_importSales.xlsx")


# =========================================================
# GENERATE FOR GROUP 2 (5-12 Temmuz, Levent)
# =========================================================
print("\n--- Generating Group 2 Files (5-12 Temmuz, Levent) ---")
pax2 = parse_rooming_strict(os.path.join(f2_dir, "0507-1207 RL.xlsx"))
exc2, sales2 = parse_sales_strict(os.path.join(f2_dir, "0507_1207 tur listesi.xlsx"))
price2 = parse_fatura_handling(os.path.join(f2_dir, "05_12.07 fatura.xlsx"))

g2_info = {
    "project_code": "PRJ-BVP",
    "description": "5-12 Temmuz Levent Orta Avrupa Turu",
    "tour_code": "PVB05072026",
    "guide_name": "Levent",
    "start_date": "05.07.2026",
    "end_date": "12.07.2026"
}

wb_meta2 = create_master_data_metadata_wb(g2_info, pax2, exc2)
wb_meta2.save(os.path.join(f2_dir, "5-12_Temmuz_Levent_importMetadata.xlsx"))
print("Saved Master Data Metadata Template: 5-12_Temmuz_Levent_importMetadata.xlsx")

wb_room2 = create_rooming_manifest_wb(pax2)
wb_room2.save(os.path.join(f2_dir, "5-12_Temmuz_Levent_importrooming.xlsx"))
print("Saved Rooming Manifest Template: 5-12_Temmuz_Levent_importrooming.xlsx")

wb_sales2 = create_sales_matrix_wb(pax2, exc2, sales2, price2)
wb_sales2.save(os.path.join(f2_dir, "5-12_Temmuz_Levent_importSales.xlsx"))
print("Saved Sales Matrix Template: 5-12_Temmuz_Levent_importSales.xlsx")

print("\nREGENERATION COMPLETE WITH MATCHING TEMPLATES & FILTERED SALES!")
