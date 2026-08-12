import openpyxl
from openpyxl import Workbook
import random
import shutil
import os

def create_masterdata_excel(filename):
    wb = Workbook()
    
    # 1. Clients
    ws_clients = wb.active
    ws_clients.title = 'Clients'
    ws_clients.append(['Name', 'Location', 'AvatarUrl', 'ContactName', 'ContactRole', 'Email', 'Phone'])
    ws_clients.append(['Mock Agency A', 'London', '', 'John Doe', 'Manager', 'john@agency-a.com', '+44 123 4567'])
    ws_clients.append(['Mock Agency B', 'Berlin', '', 'Jane Smith', 'Director', 'jane@agency-b.com', '+49 123 4567'])

    # 2. Hotels
    ws_hotels = wb.create_sheet('Hotels')
    ws_hotels.append(['Name', 'Location', 'StarRating', 'SingleRate', 'DoubleRate', 'TwinRate', 'TripleRate', 'ContactName', 'ContactRole', 'Email', 'Phone'])
    ws_hotels.append(['Budapest Grand Hotel', 'Budapest', 4, 150, 200, 200, 250, 'Hans Peter', 'Manager', 'budapest@example.com', '+36 123'])
    ws_hotels.append(['Vienna Imperial', 'Vienna', 5, 200, 250, 250, 300, 'Franz Josef', 'Director', 'vienna@example.com', '+43 123'])
    ws_hotels.append(['Prague Castle View', 'Prague', 4, 120, 180, 180, 220, 'Jan Novak', 'Owner', 'prague@example.com', '+420 123'])

    # 3. Guides
    ws_guides = wb.create_sheet('Guides')
    ws_guides.append(['Name', 'Language', 'PhoneNumber', 'DailyRate'])
    ws_guides.append(['Anna Kovacs', 'English, Hungarian', '+36 30 123 4567', 150])
    ws_guides.append(['Jan Novak', 'English, Czech, German', '+420 602 123 456', 160])

    # 4. Transport Companies
    ws_transport = wb.create_sheet('TransportCompanies')
    ws_transport.append(['Name', 'FleetSize', 'DailyRate', 'ContactName', 'ContactRole', 'Email', 'Phone'])
    ws_transport.append(['EuroBus Transfers', 15, 500, 'Alex Bus', 'Owner', 'eurobus@example.com', '+49 987'])

    # 5. Drivers
    ws_drivers = wb.create_sheet('Drivers')
    ws_drivers.append(['Name', 'PhoneNumber', 'DailyRate', 'TransportCompanyName'])
    ws_drivers.append(['Klaus Schmidt', '+49 170 1234567', 100, 'EuroBus Transfers'])
    ws_drivers.append(['Hans Meyer', '+49 171 7654321', 100, 'EuroBus Transfers'])
    ws_drivers.append(['Peter Wagner', '+49 172 5555555', 100, 'EuroBus Transfers'])

    # 6. Excursions
    ws_excursions = wb.create_sheet('Excursions')
    ws_excursions.append(['Name', 'Type', 'Price', 'SalePrice', 'VendorName'])
    ws_excursions.append(['City A Visit half day', 'Tour', 30, 50, ''])
    ws_excursions.append(['Restaurant XYZ, Evening', 'Dinner', 45, 75, ''])
    ws_excursions.append(['Prague Castle Tour', 'Tour', 25, 40, ''])
    ws_excursions.append(['Danube River Cruise', 'Activity', 35, 60, ''])

    # 7. Projects
    ws_projects = wb.create_sheet('Projects')
    ws_projects.append(['ProjectCode', 'Client', 'StartDate', 'EndDate', 'Description', 'ApproxBudget', 'BaseCurrency'])
    ws_projects.append(['PRJ-BVP1', 'Mock Agency A', '2026-08-01', '2026-08-31', 'Mock Agency A - BVP Summer Project', 50000, 'EUR'])
    ws_projects.append(['PRJ-BVP2', 'Mock Agency B', '2026-09-01', '2026-09-30', 'Mock Agency B - BVP Autumn Project', 60000, 'EUR'])

    wb.save(filename)

if __name__ == '__main__':
    base_file = 'MasterData_Import_Template.xlsx'
    create_masterdata_excel(base_file)
    
    api_path = os.path.join('wwwroot', 'templates', base_file)
    crm_path = os.path.join('..', '..', 'Uno_CRM', 'public', 'templates', base_file)
    
    os.makedirs(os.path.dirname(api_path), exist_ok=True)
    shutil.copy(base_file, api_path)
    
    os.makedirs(os.path.dirname(crm_path), exist_ok=True)
    shutil.copy(base_file, crm_path)
    
    print("Successfully generated and distributed Master Data Excel template.")
