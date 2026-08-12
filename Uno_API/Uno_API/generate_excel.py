import openpyxl
from openpyxl import Workbook
import random
import shutil
import os
from datetime import datetime, timedelta

def create_mock_excel(filename):
    wb = Workbook()
    
    # 1. Clients
    ws_clients = wb.active
    ws_clients.title = 'Clients'
    ws_clients.append(['Name', 'Location'])
    ws_clients.append(['Mock Agency A', 'London'])
    ws_clients.append(['Mock Agency B', 'Berlin'])

    # 2. Hotels
    ws_hotels = wb.create_sheet('Hotels')
    ws_hotels.append(['Name', 'Location', 'StarRating', 'SingleRate', 'DoubleRate', 'TripleRate'])
    ws_hotels.append(['Budapest Grand Hotel', 'Budapest', 4, 150, 200, 250])
    ws_hotels.append(['Vienna Imperial', 'Vienna', 5, 200, 250, 300])
    ws_hotels.append(['Prague Castle View', 'Prague', 4, 120, 180, 220])

    # 3. Guides
    ws_guides = wb.create_sheet('Guides')
    ws_guides.append(['Name', 'PhoneNumber', 'DailyRate', 'Languages'])
    ws_guides.append(['Anna Kovacs', '+36 30 123 4567', 150, 'English, Hungarian'])
    ws_guides.append(['Jan Novak', '+420 602 123 456', 160, 'English, Czech, German'])

    # 4. Transport Companies
    ws_transport = wb.create_sheet('TransportCompanies')
    ws_transport.append(['Name', 'ContactInfo', 'FleetSize', 'DailyRate'])
    ws_transport.append(['EuroBus Transfers', 'eurobus@example.com', 15, 500])

    # 5. Drivers
    ws_drivers = wb.create_sheet('Drivers')
    ws_drivers.append(['Name', 'PhoneNumber', 'DailyRate'])
    ws_drivers.append(['Klaus Schmidt', '+49 170 1234567', 100])
    ws_drivers.append(['Hans Meyer', '+49 171 7654321', 100])
    ws_drivers.append(['Peter Wagner', '+49 172 5555555', 100])

    # 6. Excursions
    ws_excursions = wb.create_sheet('Excursions')
    ws_excursions.append(['Name', 'Type', 'Price', 'SalePrice'])
    ws_excursions.append(['City A Visit half day', 'Tour', 30, 50])
    ws_excursions.append(['Restaurant XYZ, Evening', 'Dinner', 45, 75])
    ws_excursions.append(['Prague Castle Tour', 'Tour', 25, 40])
    ws_excursions.append(['Danube River Cruise', 'Activity', 35, 60])

    # 7. Projects
    ws_projects = wb.create_sheet('Projects')
    ws_projects.append(['ProjectCode', 'Description', 'StartDate', 'EndDate', 'ApproxBudget', 'BaseCurrency', 'ClientName'])
    ws_projects.append(['PRJ-BVP1', 'Mock Agency A - BVP Summer Project', '2026-08-01', '2026-08-31', 50000, 'EUR', 'Mock Agency A'])
    ws_projects.append(['PRJ-BVP2', 'Mock Agency B - BVP Autumn Project', '2026-09-01', '2026-09-30', 60000, 'EUR', 'Mock Agency B'])

    # 8. Tours
    ws_tours = wb.create_sheet('Tours')
    ws_tours.append(['ProjectCode', 'TourCode', 'Destination', 'StartDate', 'EndDate', 'Pax', 'Adults', 'Children', 'Infants'])
    tours = [
        ('PRJ-BVP1', 'BVP-1-001', 'Central Europe', '2026-08-10', '2026-08-17', 39, 30, 7, 2),
        ('PRJ-BVP1', 'BVP-1-002', 'Central Europe', '2026-08-15', '2026-08-22', 42, 35, 5, 2),
        ('PRJ-BVP1', 'BVP-1-003', 'Central Europe', '2026-08-20', '2026-08-27', 34, 28, 4, 2),
        ('PRJ-BVP2', 'BVP-2-001', 'Central Europe', '2026-09-10', '2026-09-17', 39, 32, 6, 1),
        ('PRJ-BVP2', 'BVP-2-002', 'Central Europe', '2026-09-15', '2026-09-22', 42, 36, 4, 2),
        ('PRJ-BVP2', 'BVP-2-003', 'Central Europe', '2026-09-20', '2026-09-27', 34, 30, 4, 0),
    ]
    for t in tours:
        ws_tours.append(t)

    # 9. TourServices
    ws_services = wb.create_sheet('TourServices')
    ws_services.append(['TourCode', 'ServiceType', 'HotelName', 'RoomType', 'RoomCount', 'TotalNights', 'StartDate', 'EndDate', 'IsRevenue', 'UnitPrice', 'Description'])
    
    for i, t in enumerate(tours):
        tcode = t[1]
        pax = t[5]
        start_date_str = t[3]
        start = datetime.strptime(start_date_str, '%Y-%m-%d')
        
        # Calculate mathematical room breakdown:
        # e.g., 39 pax -> 15 Double (30), 3 Triple (9)
        triples = pax % 2
        remaining = pax - (triples * 3)
        doubles = remaining // 2
        
        # Budapest (2 nights)
        ws_services.append([tcode, 'Hotel', 'Budapest Grand Hotel', 'Double', doubles, 2, start.strftime('%Y-%m-%d'), (start + timedelta(days=2)).strftime('%Y-%m-%d'), False, 200, ''])
        if triples > 0:
            ws_services.append([tcode, 'Hotel', 'Budapest Grand Hotel', 'Triple', triples, 2, start.strftime('%Y-%m-%d'), (start + timedelta(days=2)).strftime('%Y-%m-%d'), False, 250, ''])
        
        # Vienna (3 nights)
        start = start + timedelta(days=2)
        ws_services.append([tcode, 'Hotel', 'Vienna Imperial', 'Double', doubles, 3, start.strftime('%Y-%m-%d'), (start + timedelta(days=3)).strftime('%Y-%m-%d'), False, 250, ''])
        if triples > 0:
            ws_services.append([tcode, 'Hotel', 'Vienna Imperial', 'Triple', triples, 3, start.strftime('%Y-%m-%d'), (start + timedelta(days=3)).strftime('%Y-%m-%d'), False, 300, ''])

        # Prague (2 nights)
        start = start + timedelta(days=3)
        ws_services.append([tcode, 'Hotel', 'Prague Castle View', 'Double', doubles, 2, start.strftime('%Y-%m-%d'), (start + timedelta(days=2)).strftime('%Y-%m-%d'), False, 180, ''])
        if triples > 0:
            ws_services.append([tcode, 'Hotel', 'Prague Castle View', 'Triple', triples, 2, start.strftime('%Y-%m-%d'), (start + timedelta(days=2)).strftime('%Y-%m-%d'), False, 220, ''])

        # Taxes: First two tours get Revenue, next two get Expense
        if i < 2:
            ws_services.append([tcode, 'Tax', '', '', pax, 1, t[3], t[4], True, 100, 'Tax Per Pax (Revenue)'])
        elif i < 4:
            ws_services.append([tcode, 'Tax', '', '', pax, 1, t[3], t[4], False, 5, 'City Tax (Expense)'])
            ws_services.append([tcode, 'Tax', '', '', pax, 1, t[3], t[4], False, 8, 'City Tax (Expense)'])

    # 10. Bookings (Passenger Manifest)
    ws_bookings = wb.create_sheet('Bookings')
    
    first_names = ['Ahmet', 'Mehmet', 'Ayse', 'Fatma', 'Ali', 'Veli', 'Hasan', 'Huseyin', 'Mustafa', 'Kemal']
    last_names = ['Yilmaz', 'Kaya', 'Demir', 'Sahin', 'Celik', 'Yildiz', 'Yildirim', 'Ozturk', 'Aydin', 'Ozdemir']
    
    for t in tours:
        tcode = t[1]
        
        # Group Header Row
        ws_bookings.append([
            tcode, "Oda", "Pax", "Yolcu Adı", "Yolcu Soyadı", "Cinsiyet", "Skipped", 
            "T.C. Kimlik No", "Doğum Tarihi", "Telefon", "Pasaport No", "Pasaport type", "Visa No", "Address"
        ])
        
        # 5 Data Rows per tour
        for j in range(5):
            fname = random.choice(first_names)
            lname = random.choice(last_names)
            gender = 'MR' if j % 2 == 0 else 'MRS'
            tc = f"{random.randint(10000000000, 99999999999)}"
            dob = f"{random.randint(1,28):02d}/{random.randint(1,12):02d}/{random.randint(1960,2005)}"
            phone = f"90 5{random.randint(10,99)} {random.randint(100,999)} {random.randint(10,99)} {random.randint(10,99)}"
            pass_no = f"{random.randint(100000000,999999999)}"
            pass_type = "Green" if j % 3 == 0 else "Red"
            visa_no = f"{random.randint(100000,999999)}" if pass_type == "Red" else ""
            room = "DBL ROOM" if j % 2 == 0 else "SGL ROOM"
            pax_c = 2 if room == "DBL ROOM" else 1
            address = "Istanbul, Turkey"
            
            ws_bookings.append([
                "", room, pax_c, fname, lname, gender, "", tc, dob, phone, pass_no, pass_type, visa_no, address
            ])

    wb.save(filename)

if __name__ == '__main__':
    base_file = 'uno_import_template.xlsx'
    create_mock_excel(base_file)
    
    api_path = os.path.join('wwwroot', 'templates', base_file)
    crm_path = os.path.join('..', '..', 'Uno_CRM', 'public', 'templates', base_file)
    
    os.makedirs(os.path.dirname(api_path), exist_ok=True)
    shutil.copy(base_file, api_path)
    
    os.makedirs(os.path.dirname(crm_path), exist_ok=True)
    shutil.copy(base_file, crm_path)
    
    print("Successfully generated and distributed Excel templates.")
